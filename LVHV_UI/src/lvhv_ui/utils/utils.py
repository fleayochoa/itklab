from enum import Enum

class PloterStatus(Enum):
    RUNNING = 0
    STOPPED = 1
    PAUSED = 2

import numpy as np
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook


class XLSXLoader:
    def __init__(self):
        self.file_path = "test.xlsx"
        self.wb = None
        self.ws = None
        self.df = None
        self.COL = {
            "FlexID": 1,
            "PP Board Used": 5,
            "VIN+": 6,
            "VIN-": 7,
            "GND+": 8,
            "GND-": 9,
            "C leak test (mV)": 10,
            "NTC reading (V)": 11
        }

    def set_file_path(self, path):
        self.file_path = path

    def load_data(self):
        try:
            self.wb = load_workbook(self.file_path)
            self.ws = self.wb["LVHV_copy"]
            self.df = pd.read_excel(self.file_path, sheet_name="LVHV_copy")
        except Exception as e:
            raise ValueError("Could not load Excel file.")
    
    def get_ids_with_empty_ntc(self):
        if self.df is None:
            raise ValueError("Data not loaded. Call load_data() first.")
        ids_con_ntc_vacio = self.df.loc[self.df["NTC reading (V)"].isna(), "FlexID"].tolist()
        return ids_con_ntc_vacio

    def overwrite_xlsx(self, data_str, flexID, pogo):
        if self.df is None:
            raise ValueError("Data not loaded. Call load_data() first.")
        arr_data = data_str.split("_")
        rows_flex0 = self.df.index[self.df["FlexID"] == flexID[0]]
        rows_flex1 = self.df.index[self.df["FlexID"] == flexID[1]]
        for i in rows_flex0:
            r = i+2  # +2 por el header y el index base 0

            self.ws.cell(r, self.COL["PP Board Used"], int(pogo[0]))
            #celda = self.ws.cell(r, self.COL["PP Board Used"])
            #celda.value = float(pogo[0])
            #celda.hyperlink = None
            #celda.font = celda.font.copy(underline=None)
            self.ws.cell(r, self.COL["VIN+"], float(arr_data[0]))
            self.ws.cell(r, self.COL["VIN-"], float(arr_data[1]))
            self.ws.cell(r, self.COL["GND+"], float(arr_data[2]))
            self.ws.cell(r, self.COL["GND-"], float(arr_data[3]))
            self.ws.cell(r, self.COL["C leak test (mV)"], float(arr_data[4]))
            self.ws.cell(r, self.COL["NTC reading (V)"], float(arr_data[5]))

        for i in rows_flex1:
            r = i+2  # +2 por el header y el index base 0

            self.ws.cell(r, self.COL["PP Board Used"], int(pogo[1]))
            self.ws.cell(r, self.COL["VIN+"], float(arr_data[6]))
            self.ws.cell(r, self.COL["VIN-"], float(arr_data[7]))
            self.ws.cell(r, self.COL["GND+"], float(arr_data[8]))
            self.ws.cell(r, self.COL["GND-"], float(arr_data[9]))
            self.ws.cell(r, self.COL["C leak test (mV)"], float(arr_data[10]))
            self.ws.cell(r, self.COL["NTC reading (V)"], float(arr_data[11]))

        print(flexID, pogo, data_str)
        self.wb.save(self.file_path)